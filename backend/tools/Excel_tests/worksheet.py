"""Worksheet abstraction exposing normalized cell access utilities."""

from __future__ import annotations

import logging
from typing import Iterator

from openpyxl.worksheet.worksheet import Worksheet

from .cell import CellValue, ExcelCell


logger = logging.getLogger(__name__)


class ExcelWorksheet:
    """Provides typed access to worksheet cells and iteration helpers."""

    def __init__(self, worksheet: Worksheet) -> None:
        """Wraps an openpyxl worksheet instance.

        Args:
            worksheet: Source worksheet object.
        """
        self._worksheet = worksheet

    @property
    def name(self) -> str:
        """Returns worksheet display name."""
        return self._worksheet.title

    @property
    def max_row(self) -> int:
        """Returns index of last populated row reported by openpyxl."""
        return self._worksheet.max_row

    @property
    def max_column(self) -> int:
        """Returns index of last populated column reported by openpyxl."""
        return self._worksheet.max_column

    def get_cell(self, row: int, column: int) -> ExcelCell:
        """Returns normalized cell representation for given coordinates.

        Args:
            row: One-based row index.
            column: One-based column index.

        Returns:
            Normalized ExcelCell payload.

        Raises:
            ValueError: When row or column index is invalid.
        """
        if row < 1 or column < 1:
            raise ValueError("Row and column indexes must be greater than 0")

        raw_cell = self._worksheet.cell(row=row, column=column)
        excel_cell = ExcelCell(
            row=row,
            column=column,
            coordinate=raw_cell.coordinate,
            value=self._normalize_value(raw_cell.value),
            fill_color=self._extract_fill_color(raw_cell),
            fill_pattern=self._extract_fill_pattern(raw_cell),
            font_color=self._extract_font_color(raw_cell),
            has_diagonal_down=self._has_diagonal_down(raw_cell),
            is_merged=self._is_merged_cell(raw_cell.coordinate),
        )
        return excel_cell

    def iter_rows(self) -> Iterator[list[ExcelCell]]:
        """Iterates over worksheet rows as normalized cell lists."""
        for row_idx in range(1, self.max_row + 1):
            yield [
                self.get_cell(row=row_idx, column=column_idx)
                for column_idx in range(1, self.max_column + 1)
            ]

    def iter_cells(self) -> Iterator[ExcelCell]:
        """Iterates over all worksheet cells in row-major order."""
        for row in self.iter_rows():
            for cell in row:
                yield cell

    def _is_merged_cell(self, coordinate: str) -> bool:
        for merged_range in self._worksheet.merged_cells.ranges:
            if coordinate in merged_range:
                return True
        return False

    def _extract_fill_color(self, raw_cell) -> str | None:
        fill = raw_cell.fill
        if fill is None or fill.fill_type is None:
            return None
        return self._extract_openpyxl_color(raw_cell.fill.start_color)

    def _extract_font_color(self, raw_cell) -> str | None:
        font = raw_cell.font
        if font is None or font.color is None:
            return None
        return self._extract_openpyxl_color(font.color)

    def _extract_fill_pattern(self, raw_cell) -> str | None:
        fill = raw_cell.fill
        if fill is None:
            return None
        pattern = getattr(fill, "patternType", None)
        if pattern is None:
            return None
        return str(pattern)

    def _has_diagonal_down(self, raw_cell) -> bool:
        border = raw_cell.border
        if border is None:
            return False
        return bool(getattr(border, "diagonalDown", False))

    def _extract_openpyxl_color(self, color) -> str | None:
        if color is None:
            return None

        color_type = getattr(color, "type", None)
        if color_type == "rgb" and getattr(color, "rgb", None):
            return str(color.rgb).upper()

        if color_type == "indexed" and getattr(color, "indexed", None) is not None:
            return f"INDEXED:{color.indexed}"

        if color_type == "theme" and getattr(color, "theme", None) is not None:
            tint = getattr(color, "tint", 0)
            return f"THEME:{color.theme}:TINT:{tint}"

        return None

    def _normalize_value(self, value) -> CellValue:
        if value is None:
            return None
        if isinstance(value, (str, int, float)):
            return value

        normalized = str(value)
        logger.debug("Converted unsupported cell value type to string: %s", type(value))
        return normalized
