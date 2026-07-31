"""Workbook abstraction for validated loading and worksheet access."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .exceptions import InvalidWorkbookFormatError, WorkbookNotFoundError, WorksheetNotFoundError
from .worksheet import ExcelWorksheet


logger = logging.getLogger(__name__)


class ExcelWorkbook:
    """Provides safe access to workbook metadata and worksheet objects."""

    def __init__(self, workbook: Workbook, source_path: Path) -> None:
        """Wraps an openpyxl workbook loaded from a source path.

        Args:
            workbook: Loaded openpyxl workbook instance.
            source_path: Filesystem path of the workbook source file.
        """
        self._workbook = workbook
        self._source_path = source_path

    @classmethod
    def load(cls, path: str | Path) -> "ExcelWorkbook":
        """Loads an .xlsx workbook and validates source path.

        Args:
            path: Path to workbook file.

        Returns:
            Loaded workbook wrapper.

        Raises:
            WorkbookNotFoundError: When file does not exist.
            InvalidWorkbookFormatError: When extension is not .xlsx.
        """
        workbook_path = Path(path)
        logger.info("Loading Excel workbook from %s", workbook_path)

        if not workbook_path.exists() or not workbook_path.is_file():
            logger.error("Workbook path does not exist: %s", workbook_path)
            raise WorkbookNotFoundError(f"Workbook not found: {workbook_path}")

        if workbook_path.suffix.lower() != ".xlsx":
            logger.error("Unsupported workbook format for path: %s", workbook_path)
            raise InvalidWorkbookFormatError(
                f"Unsupported workbook format: {workbook_path.suffix}. Only .xlsx is supported"
            )

        workbook = load_workbook(filename=workbook_path, data_only=True)
        return cls(workbook=workbook, source_path=workbook_path)

    def get_sheet_names(self) -> list[str]:
        """Returns workbook sheet names in their original order."""
        return list(self._workbook.sheetnames)

    def get_sheet(self, name: str) -> ExcelWorksheet:
        """Returns worksheet wrapper for the given sheet name.

        Args:
            name: Worksheet name.

        Returns:
            Wrapped worksheet object.

        Raises:
            WorksheetNotFoundError: When sheet name is not present.
        """
        if name not in self._workbook.sheetnames:
            raise WorksheetNotFoundError(f"Worksheet not found: {name}")
        return ExcelWorksheet(self._workbook[name])

    def get_all_sheets(self) -> list[ExcelWorksheet]:
        """Returns wrappers for all worksheets in the workbook."""
        return [self.get_sheet(name) for name in self.get_sheet_names()]
